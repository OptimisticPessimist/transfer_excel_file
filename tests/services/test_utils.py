from pathlib import Path

from pytest import mark
from openpyxl import load_workbook

from src.services.utils import *


TEST_DATA_DIR = Path("./tests/data")


class TestGetMergedCellValue:

    @mark.parametrize("sheet, cell, expected", (
            ["Sheet1", "B2", "a"],
            ["Sheet1", "B1", 1],
            ["Sheet1", "A6", 2]
    ))
    def test_get_merged_cell_value(self, sheet, cell, expected):
        wb = load_workbook(TEST_DATA_DIR / "test.xlsx", keep_vba=True, keep_links=True)
        ws = wb[sheet]
        cell = ws[cell]
        actual = fetch_merged_cell_value(ws, cell)
        assert actual == expected
